from __future__ import annotations

import argparse
import csv
import json
import mimetypes
import time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from isb_igraph.environment import assert_runtime_compatibility
from isb_igraph.pipeline import run_pipeline
from isb_igraph.runtime import (
    build_pipeline_config_dict,
    ensure_runtime_dirs,
    jobs_db_path,
    pipeline_config_from_dict,
)

from services.api.store import (
    JOB_STATUS_CANCELLED,
    JobRecord,
    JobStore,
)


class JobCancelledError(RuntimeError):
    pass


def _json_loads_safe(value: str | None) -> dict[str, Any]:
    if not value:
        return {}
    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def _xlsx_to_csv(xlsx_path: Path, csv_path: Path) -> Path:
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    try:
        sheet = wb[wb.sheetnames[0]]
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
    finally:
        wb.close()
    return csv_path


def _parse_gcs_uri(gcs_uri: str) -> tuple[str, str]:
    if not gcs_uri.startswith("gs://"):
        raise ValueError(f"Invalid GCS URI: {gcs_uri}")
    remainder = gcs_uri[len("gs://") :]
    bucket, sep, blob = remainder.partition("/")
    if not bucket or not sep or not blob:
        raise ValueError(f"Invalid GCS URI: {gcs_uri}")
    return bucket, blob


def _download_gcs_to_local(gcs_uri: str, output_dir: Path) -> Path:
    try:
        from google.cloud import storage  # type: ignore
    except Exception as exc:
        raise RuntimeError("google-cloud-storage is required for gs:// inputs") from exc

    bucket_name, blob_name = _parse_gcs_uri(gcs_uri)
    output_dir.mkdir(parents=True, exist_ok=True)

    blob_file_name = Path(blob_name).name or "input_from_gcs.csv"
    local_path = output_dir / blob_file_name

    client = storage.Client()
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(blob_name)
    blob.download_to_filename(str(local_path))

    if not local_path.exists() or not local_path.is_file():
        raise FileNotFoundError(f"Downloaded file missing after GCS fetch: {local_path}")
    return local_path


def _collect_artifacts(output_files: dict[str, Path]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for artifact_name, file_path in output_files.items():
        path = Path(file_path)
        if not path.exists() or not path.is_file():
            continue
        mime_type = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
        out.append(
            {
                "artifact_name": artifact_name,
                "file_path": str(path),
                "mime_type": mime_type,
                "size_bytes": int(path.stat().st_size),
            }
        )
    return out


def _ensure_not_cancelled(store: JobStore, job_id: str) -> None:
    status = store.get_job_status(job_id)
    if status == JOB_STATUS_CANCELLED:
        raise JobCancelledError("Job cancelled")


def _run_one_job(store: JobStore, job: JobRecord) -> None:
    options = _json_loads_safe(job.options_json)
    output_dir = Path(job.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    input_ref = str(job.input_path)

    if input_ref.startswith("gs://"):
        store.add_event(job.job_id, level="info", message=f"Downloading input from GCS: {input_ref}", progress=0.01)
        input_path = _download_gcs_to_local(input_ref, output_dir / "input_from_gcs")
        store.add_event(job.job_id, level="info", message=f"Downloaded GCS input to {input_path}", progress=0.02)
    else:
        input_path = Path(input_ref)
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_path}")

    resolved_input = input_path
    if input_path.suffix.lower() == ".xlsx":
        store.add_event(job.job_id, level="info", message="Converting XLSX to CSV", progress=0.03)
        converted_path = output_dir / "input_from_xlsx.csv"
        resolved_input = _xlsx_to_csv(input_path, converted_path)

    config_dict = build_pipeline_config_dict(
        input_csv=resolved_input,
        output_dir=output_dir,
        options=options,
    )
    config = pipeline_config_from_dict(config_dict)

    def on_progress(message: str, fraction: float) -> None:
        _ensure_not_cancelled(store, job.job_id)
        fraction = max(0.0, min(1.0, float(fraction)))
        store.update_job_progress(job.job_id, stage=message, progress=fraction)
        store.add_event(job.job_id, level="info", message=message, progress=fraction)

    store.add_event(job.job_id, level="info", message="Job started", progress=0.0)
    result = run_pipeline(config, progress_fn=on_progress)

    artifacts = _collect_artifacts(result.output_files)
    store.replace_artifacts(job.job_id, artifacts)

    run_metadata = {
        "graph_summary": result.graph_summary,
        "validation_report": result.validation_report,
        "qa_report": result.qa_report,
        "profile_records": result.profile_records,
        "output_files": {k: str(v) for k, v in result.output_files.items()},
    }
    store.set_job_completed(job.job_id, run_metadata=run_metadata)
    store.add_event(job.job_id, level="info", message="Job completed", progress=1.0)


def run_worker(*, poll_interval_seconds: float = 2.0, once: bool = False) -> None:
    ensure_runtime_dirs()
    store = JobStore(jobs_db_path())

    while True:
        job = store.claim_next_queued_job()
        if job is None:
            if once:
                return
            time.sleep(max(0.25, poll_interval_seconds))
            continue

        try:
            _run_one_job(store, job)
        except JobCancelledError:
            store.cancel_job(job.job_id)
            store.add_event(job.job_id, level="warning", message="Job cancelled during execution", progress=None)
        except Exception as exc:
            # If user cancelled while exception was being raised, keep cancelled state.
            current_status = store.get_job_status(job.job_id)
            if current_status != JOB_STATUS_CANCELLED:
                store.set_job_failed(job.job_id, error_message=str(exc))
                store.add_event(job.job_id, level="error", message=f"Job failed: {exc}", progress=None)

        if once:
            return


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Run ISB iGraph async worker")
    parser.add_argument("--poll-interval", type=float, default=2.0)
    parser.add_argument("--once", action="store_true", help="Process at most one queued job and exit")
    return parser


def main() -> None:
    assert_runtime_compatibility()
    args = _build_parser().parse_args()
    run_worker(poll_interval_seconds=float(args.poll_interval), once=bool(args.once))


if __name__ == "__main__":
    main()
